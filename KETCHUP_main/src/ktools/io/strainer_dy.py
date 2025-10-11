"""
Strainer file import.

Importation of spreadsheet-based strainer file data for model and mechanism data.
Currently only supports experimental data import.

"""

import pandas as pd
import openpyxl
from .dataframes import parse_strainer_dy_data_df

def read_strainer_data_xlsx(filename_data: str, data_headers: str or dict = None,
                            data_type: str = 'dynamic', debug: bool = False) -> pd.DataFrame:
    """
    Reads and processes a strainer spreadsheet file containing experimental data.

    Parameters
    ----------
    filename_data : str
        The filename of the strainer spreadsheet file containing experimental data.
    data_headers : str or dict, optional
        The header information to use to identify the location and types of data in the
        strainer spreadsheet file.
        Defaults to None.
    data_type : str
        Type of the data in the file ('static' or 'dynamic').
        Defaults to 'static'.
    debug : bool, optional
        If True, prints debugging messages. Defaults to False.

    Returns
    -------
    pd.DataFrame
        A pandas DataFrame containing the initial strainer experimental data.
    """

    if data_type == 'static':
        # for right now, only supports dynamic data
        import warnings
        warnings.warn("Warning: strainer files only currently support dynamic data.")
    elif data_type == 'dynamic':
        if data_headers == None:
            import warnings
            warnings.warn("Warning: strainer requires externally supplied headers of dynamic data in json format.")
            return
        else:
            import openpyxl
            workbook = openpyxl.load_workbook(filename=filename_data, data_only=True)
            flag_first_sheet = True
            for sheetname in workbook.sheetnames:
                #print(sheetname)
                dy_df_dat = _xlsx_to_df(filename_data, sheetname, debug=debug)

                if debug: print(sheetname, dy_df_dat)

                # move header down
                dy_df_dat.loc[-1] = dy_df_dat.columns.to_list()
                dy_df_dat = dy_df_dat.sort_index().reset_index(drop=True)
                dy_df_dat.columns = range(dy_df_dat.shape[1])

                orig_names = ["experiment ID"] + dy_df_dat.loc[len(data_headers['t_0'])].tolist() + \
                             dy_df_dat[:len(data_headers['t_0'])][0].tolist()
                if debug: print(orig_names)

                if flag_first_sheet:
                    flag_first_sheet = False
                    # expand header field "status" to be full word if not already
                    expanded_status = []
                    for item in data_headers["status"]:
                        if item == "i":
                            expanded_status.append("independent")
                        elif item == "g":
                            expanded_status.append("ignore")
                        elif item == "d":
                            expanded_status.append("dependent")
                        else:
                            expanded_status.append(item)
                    data_headers["status"] = expanded_status

                    # create new data frame
                    new_columns = ["experiment ID", "Time"] + [f"Values[{m}]" for m in data_headers["time"]] + \
                                  [f"Values[{m1}]_0" for m1 in data_headers["t_0"]]
                    header_1 = ["id", "t"] + [t for t in data_headers["type"][-len(data_headers["time"]):]] + \
                               [t for t in data_headers["type"][:len(data_headers["t_0"])]]
                    header_2 = ["meta", "time"] + [t for t in data_headers["status"][-len(data_headers["time"]):]] + \
                               [t for t in data_headers["status"][:len(data_headers["t_0"])]]
                    # print (header_2)

                    df_composite = pd.DataFrame(columns=new_columns, data=[header_1, header_2])
                    # print (new_columns)

                # slice off initial conditions
                df_tmp_t0 = dy_df_dat[:len(data_headers['t_0'])]
                # slice off dynamic portion
                df_tmp_dynamic = dy_df_dat[len(data_headers['t_0']) + 1:]

                # insert experiment ID
                df_tmp_A = df_tmp_dynamic.copy()
                df_tmp_A.insert(0, -1, sheetname)
                df_tmp_A = df_tmp_A.reset_index(drop=True)

                # transpose tmp_t0 to prepare for merge
                df_tmp_t0t = df_tmp_t0.transpose()
                df_tmp_t0t = df_tmp_t0t.drop(df_tmp_t0t.index[0]).reset_index(drop=True)
                df_tmp_t0t.columns = range(df_tmp_A.columns[-1] + 1, df_tmp_A.columns[-1] + 1 + df_tmp_t0t.shape[1])

                # merge into composite list
                df_tmp = (df_tmp_A.join(df_tmp_t0t))
                df_tmp.columns = new_columns
                df_composite = pd.concat([df_composite, df_tmp])
                # display (tmp_t0t)
    else:
        import warnings
        warnings.warn("Warning: Unsupported data type.")

    df_composite = df_composite.reset_index(drop=True)

    df_data = parse_strainer_dy_data_df(df_composite, debug=debug)

    return df_data

# TODO: following is same as the one in kfit_ss.py, so consider creating a unified function in utils
def _xlsx_to_df(filename_workbook, sheet_name=None, debug=False):
    try:
        df = pd.DataFrame()
        if debug:
            print(f"Beginning to read file {filename_workbook}")
        # Note that 'data_only=True' reads the results for formulae
        #   instead of the formulae (default)
        workbook = openpyxl.load_workbook(filename=filename_workbook, data_only=True)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        elif sheet_name == None:
            sheet = workbook[workbook.sheetnames[0]]
        if debug:
            print(f"Spreadsheet name: {filename_workbook}")
            print(f"Worksheet names: {workbook.sheetnames}")
            print(f"The title of the current Worksheet is: {sheet.title}")
            print(f"Cells that contain data: {sheet.calculate_dimension()}")
        data = sheet.values
        cols = next(data)[:]
        data = list(data)
        df = pd.DataFrame(data, columns=cols)
        return df
    except Exception:
        if sheet_name:
            print(f"Error reading {filename_workbook}, {sheet_name}")
        else:
            print(f"Error reading {filename_workbook}")
        return None
