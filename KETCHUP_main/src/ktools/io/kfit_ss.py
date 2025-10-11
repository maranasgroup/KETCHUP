"""
K-FIT spreadsheet import.

Importation of K-FIT spreadsheets for model and mechanism data.

"""

import cobra
import pandas as pd
import openpyxl
import os
from os.path import join
from .dataframes import parse_kfit_model_df
from .dataframes import parse_kfit_mech_df
from .dataframes import parse_kfit_ss_data_df
import cobra.core.model

def read_kfit_spreadsheets_xlsx(filename_model: str, filename_mechanism: str,
                    met_sheet: str = 'Metabolites', rxn_sheet = 'Reactions',
                    mech_sheet: str = None, debug: bool = False) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Read in the model K-FIT spreadsheets. Note that the default sheet
    names might need to be changed for each bespoke model.

    Parameters
    ----------
    filename_model : str
        The filename of  the K-FIT model spreadsheet file.
    filename_mechanism : str
        The filename of  the  K-FIT mechanism spreadsheet file.
    met_sheet : str, optional
        The name of the sheet containing metabolite information.
        Defaults to 'Metabolites'.
    rxn_sheet : str, optional
        The name of the worksheet containing reaction information.
        Defaults to 'Reactions'.
    mech_sheet : str, optional
        The name of the worksheet containing mechanism information.
        Defaults to None.
    debug : bool, optional
        If True, prints debugging messages.
        Defaults to False.

    Returns
    -------
    tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]
        A tuple containing three raw pandas DataFrames:
            - DataFrame containing metabolites.
            - DataFrame containing reactions.
            - DataFrame containing mechanisms.
    """

    from warnings import warn
    
    df_metabolite = _xlsx_to_df(filename_model,'Metabolites', debug=debug)
    df_reaction = _xlsx_to_df(filename_model,'Reactions', debug=debug)
    df_mechanism = _xlsx_to_df(filename_mechanism, debug=debug) # K-FIT defaults to first sheet for mechanism file
    
    if debug:
        print('All K-FIT input files processed')
    return df_metabolite, df_reaction, df_mechanism


def read_kfit_model_xlsx(filename_model, filename_mechanism,
                         met_sheet='Metabolites', rxn_sheet='Reactions',
                         mech_sheet: str = None, mech_type: str = 'elemental',
                         debug: bool = False) -> tuple[cobra.core.model.Model, pd.DataFrame]:
    """
    Read in the model K-FIT spreadsheets and create the model representations.

    Parameters
    ----------
    filename_model : str
        The filename of the K-FIT model spreadsheet file.
    filename_mechanism : str
        The filename of  the  K-FIT mechanism spreadsheet file.
    met_sheet : str, optional
        The name of the worksheet containing metabolite information.
        Defaults to 'Metabolites'.
    rxn_sheet : str, optional
        The name of the worksheet containing reaction information.
        Defaults to 'Reactions'.
    mech_sheet : str, optional
        The name of the worksheet containing mechanism information.
        Defaults to None.
    mech_type : str, optional
        The type of mechanism to use for rate laws.
        Defaults to 'elemental'.
    debug : bool, optional
        If True, prints debugging messages.
        Defaults to False.

    Returns
    -------
    tuple[cobra.core.model.Model, pd.DataFrame]
        A tuple containing
            - COBRApy representation of the reaction network.
            - pandas DataFrame containing processed mechanisms.
    """

    (ss_df_met,ss_df_rxn,ss_df_mech) = read_kfit_spreadsheets_xlsx(
                                           filename_model, filename_mechanism,
                                           met_sheet, rxn_sheet, mech_sheet, debug)

    m_model = parse_kfit_model_df(ss_df_met, ss_df_rxn, debug=debug)
    mech_df = parse_kfit_mech_df(ss_df_mech, m_model, mech_type=mech_type,
                                 debug=debug)
    return m_model, mech_df


def read_kfit_data_xlsx(filename_data: str, data_sheet: str = None,
                        debug: bool = False) -> pd.DataFrame:
    """
    Reads a K-FIT data spreadsheet file.

    Parameters
    ----------
    filename_data : str
        The filename of the K-FIT data spreadsheet file.
    data_sheet : str, optional
        The name of the worksheet to read. If None, the first sheet is read.
        Defaults to None.
    debug : bool, optional
        If True, prints debugging messages. Defaults to False.

    Returns
    -------
    pd.DataFrame
        A pandas DataFrame containing the parsed K-FIT data.
    """
    # for right now, only the first (or indicated) sheet is read and multiple calls
    #   are required to read multiple sheets, if applicable
    #
    
    df_ss_dat = _xlsx_to_df(filename_data, data_sheet, debug=debug)

    df_data = parse_kfit_ss_data_df(df_ss_dat, debug=debug)

    return df_data


def _xlsx_to_df(filename_workbook, sheet_name=None, debug=False):
    
    try:
        df = pd.DataFrame()
        if debug:
            print(f"Beginning to read file {filename_workbook}")
        # Note that 'data_only=True' reads the results for formulae
        #   instead of the formulae (default)
        workbook = openpyxl.load_workbook(filename=filename_workbook, data_only=True)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        elif sheet_name == None:
            sheet = workbook[workbook.sheetnames[0]]
        if debug:
            print(f"Spreadsheet name: {filename_workbook}")
            print(f"Worksheet names: {workbook.sheetnames}")
            print(f"The title of the current Worksheet is: {sheet.title}")
            print(f"Cells that contain data: {sheet.calculate_dimension()}")
        data = sheet.values
        cols = next(data)[:]
        data = list(data)
        df = pd.DataFrame(data, columns=cols)
        return df
    except Exception:
        if sheet_name:
            print(f"Error reading {filename_workbook}, {sheet_name}")
        else:
            print(f"Error reading {filename_workbook}")
        return None
